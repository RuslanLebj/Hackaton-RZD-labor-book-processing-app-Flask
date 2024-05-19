import json
import os
from io import BytesIO
from models import *
from PIL import Image
from openpyxl import Workbook


def convert_to_jpg(image):
    buffered = BytesIO()
    image.save(buffered, format="JPEG")
    return buffered.getvalue()


def analyze_image(data):
    return '{\"awards\": [{\"awards_information\": \"danil krasava\",\"date\": \"2222222\", \"employee_id\": 1,\"id\": 1,\"order_number\": \"2222222222\",\"seal_decryption\": \"enviro\"}],    "birth_year": "2023",\"document_issue_date\": null,\"document_issuer\": null,\"document_number\": "111111111",\"document_series\": "7502",\"document_type\": null,\"id\": 1,\"jobs\": [],\"name\": \"Данил\",\"number\": \"777777\",\"patronymic\": \"Рустамович\",\"series": \"SI-I\",\"surname": \"Мухаметшин\"}'


def generate_xlsx(employee):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Employee Data'

    ws.append(['Серия', 'Номер', 'Фамилия', 'Имя', 'Отчество', 'Год рождения', 'Документ Серия', 'Документ Номер',
               'Дата выдачи', 'Кем выдано'])
    ws.append([
        employee.series,
        employee.number,
        employee.surname,
        employee.name,
        employee.patronymic,
        employee.birth_year,
        employee.document_series,
        employee.document_number,
        employee.document_issue_date,
        employee.document_issuer
    ])

    ws2 = wb.create_sheet(title='Job Information')
    ws2.append(['Дата приема', 'Дата увольнения', 'Расшифровка печати', 'Расшифровка должности', 'Приказ'])
    for job in employee.jobs:
        ws2.append([
            job.admission_date,
            job.dismissal_date,
            job.seal_decryption,
            job.position_decryption,
            job.order_number
        ])

    ws3 = wb.create_sheet(title='Awards')
    ws3.append(['Дата', 'Расшифровка печати', 'Информация о поощрении', 'Приказ'])
    for award in employee.awards:
        ws3.append([
            award.date,
            award.seal_decryption,
            award.awards_information,
            award.order_number
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def process_single_image(image_path):
    with Image.open(image_path) as img:
        if image_path.endswith(('.png', '.svg')):
            img = img.convert('RGB')
        img_data = convert_to_jpg(img)
        model_response = analyze_image(img_data)
        return json.loads(model_response)


def process_images_from_zip(zip_folder):
    data = {}
    images_data = []
    for filename in os.listdir(zip_folder):
        if filename.endswith(('.jpg', '.jpeg', '.png', '.svg')):
            image_path = os.path.join(zip_folder, filename)
            image_data = process_single_image(image_path)
            images_data.append({
                'filename': filename,
                'data': image_data
            })
    data['images'] = images_data
    return data


def decode_json(data):
    if isinstance(data, str):
        return json.loads(data, encoding='utf-8')
    return data


def update_employee_data(employee, data):
    employee.series = data.get('series', employee.series)
    employee.number = data.get('number', employee.number)
    employee.surname = data.get('surname', employee.surname)
    employee.name = data.get('name', employee.name)
    employee.patronymic = data.get('patronymic', employee.patronymic)
    employee.birth_year = data.get('birth_year', employee.birth_year)
    employee.document_type = data.get('document_type', employee.document_type)
    employee.document_series = data.get('document_series', employee.document_series)
    employee.document_number = data.get('document_number', employee.document_number)
    employee.document_issue_date = data.get('document_issue_date', employee.document_issue_date)
    employee.document_issuer = data.get('document_issuer', employee.document_issuer)


def insert_job_information(employee_id, job_information):
    for job_info in job_information:
        job_info['employee_id'] = employee_id
        job = JobInformation(**job_info)
        db.session.add(job)


def insert_awards(employee_id, awards):
    for award_info in awards:
        award_info['employee_id'] = employee_id
        award = Award(**award_info)
        db.session.add(award)
